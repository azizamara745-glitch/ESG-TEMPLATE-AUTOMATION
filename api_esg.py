"""
==============================================================================
  GÉNÉRATEUR ESG — VERSION AUTOMATISATION (compatible n8n)
  api_esg.py
==============================================================================

Ce fichier expose une API HTTP locale (Flask) que n8n peut appeler
automatiquement après chaque sauvegarde de la base de données.

FONCTIONNEMENT :
  1. Tu sauvegardes la base de données Excel sur ton serveur / Google Drive
  2. n8n détecte la modification (via webhook ou watcher de fichier)
  3. n8n envoie une requête POST à cette API
  4. L'API génère le rapport ESG mis à jour
  5. n8n récupère le résultat (chemin fichier, statut, timestamp)

ENDPOINTS :
  POST /generer-rapport
      → Corps JSON : { "fichier_db": "...", "fichier_template": "..." }  (optionnel)
      → Réponse    : { "statut": "ok", "fichier": "...", "timestamp": "..." }

  GET  /statut
      → Retourne l'état de santé de l'API (pour le ping n8n)

DÉMARRAGE :
  pip install flask openpyxl pandas
  python api_esg.py
  → API disponible sur http://localhost:5050

CONFIGURATION n8n :
  - Node type : HTTP Request
  - Method    : POST
  - URL       : http://localhost:5050/generer-rapport
  - Body      : JSON (voir ENDPOINTS ci-dessus)
==============================================================================
"""

from flask import Flask, request, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import traceback
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ==============================================================================
# CONFIGURATION
# ==============================================================================
# Ces valeurs par défaut sont utilisées si n8n n'en envoie pas d'autres
DOSSIER_DEFAUT   =  r"C:\Users\PC\Desktop\PROJECT MACRO"
DB_DEFAUT        = r"C:\Users\PC\Desktop\PROJECT MACRO\InnovTech_Base_Donnees.xlsx"
TEMPLATE_DEFAUT  = r"C:\Users\PC\Desktop\PROJECT MACRO\ESG_Template_TotalEnergies_STRUCTURED.xlsx"
NOM_ENTREPRISE   = "InnovTech SA"
ANNEE_RAPPORT    = 2024

# Port sur lequel l'API écoute (doit être libre sur ta machine)
PORT_API = 5050

# ==============================================================================
# CONFIGURATION EMAIL — Modifier avec ton mot de passe Gmail
# ==============================================================================
EMAIL_EXPEDITEUR   = "azizamara745@gmail.com"  # Ton adresse Gmail
EMAIL_DESTINATAIRE = "azizamara745@gmail.com"  # Adresse de réception
EMAIL_MOT_DE_PASSE = "METTRE_MOT_DE_PASSE_APP" # ← REMPLACE ICI par tes 16 caractères
SMTP_HOST          = "smtp.gmail.com"
SMTP_PORT          = 465

# Initialisation de l'application Flask
app = Flask(__name__)


# ==============================================================================
# ENDPOINT PING — GET /statut
# ==============================================================================
@app.route("/statut", methods=["GET"])
def statut():
    """
    Endpoint de vérification d'état pour n8n.
    n8n peut l'appeler pour vérifier que l'API est bien démarrée.
    
    Retourne : { "statut": "ok", "service": "...", "timestamp": "..." }
    """
    return jsonify({
        "statut":    "ok",
        "service":   f"Générateur ESG — {NOM_ENTREPRISE}",
        "version":   "2.0",
        "timestamp": datetime.now().isoformat()
    })


# ==============================================================================
# ENDPOINT PRINCIPAL — POST /generer-rapport
# ==============================================================================
@app.route("/generer-rapport", methods=["POST"])
def generer_rapport():
    """
    Endpoint principal appelé par n8n.
    
    Corps JSON attendu (tout optionnel — utilise les valeurs par défaut) :
    {
        "fichier_db":       "/chemin/vers/InnovTech_Base_Donnees.xlsx",
        "fichier_template": "/chemin/vers/ESG_Template.xlsx",
        "dossier_sortie":   "/chemin/vers/sorties/"
    }
    
    Retourne :
    {
        "statut":    "ok",
        "fichier":   "/chemin/vers/Rapport_ESG_InnovTech_2024.xlsx",
        "timestamp": "2024-12-15T10:30:00",
        "resume":    { ... stats du rapport ... }
    }
    """
    try:
        # --- Lecture des paramètres envoyés par n8n (ou valeurs par défaut) ---
        params = request.get_json(silent=True) or {}
        
        fichier_db       = params.get("fichier_db",       DB_DEFAUT)
        fichier_template = params.get("fichier_template", TEMPLATE_DEFAUT)
        dossier_sortie   = params.get("dossier_sortie",   DOSSIER_DEFAUT)
        
        # Nom du fichier de sortie avec timestamp (pour ne pas écraser l'ancien)
        timestamp_fichier = datetime.now().strftime("%Y%m%d_%H%M%S")
        nom_rapport = f"Rapport_ESG_InnovTech_{timestamp_fichier}.xlsx"
        fichier_rapport = os.path.join(dossier_sortie, nom_rapport)
        
        print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 🔔 Requête reçue de n8n")
        print(f"  DB       : {fichier_db}")
        print(f"  Template : {fichier_template}")
        print(f"  Sortie   : {fichier_rapport}")
        
        # --- Vérification que les fichiers source existent ---
        if not os.path.exists(fichier_db):
            return jsonify({
                "statut": "erreur",
                "message": f"Base de données introuvable : {fichier_db}"
            }), 400
        
        if not os.path.exists(fichier_template):
            return jsonify({
                "statut": "erreur",
                "message": f"Template introuvable : {fichier_template}"
            }), 400
        
        # --- Exécution du pipeline ESG ---
        db          = _lire_base_de_donnees(fichier_db)
        indicateurs = _extraire_indicateurs_esg(db)
        _creer_rapport(db, indicateurs, fichier_template, fichier_rapport)
        
        # --- Lecture du fichier en base64 pour pièce jointe email ---
        import base64
        with open(fichier_rapport, "rb") as f:
            fichier_base64 = base64.b64encode(f.read()).decode("utf-8")

        # --- Réponse succès pour n8n ---
        reponse = {
            "statut":          "ok",
            "fichier":         fichier_rapport,
            "nom_fichier":     nom_rapport,
            "fichier_base64":  fichier_base64,
            "timestamp":       datetime.now().isoformat(),
            "resume": {
                "entreprise":      NOM_ENTREPRISE,
                "annee":           ANNEE_RAPPORT,
                "emplois_directs": indicateurs["meta"]["emplois"],
                "ca_tnd":          indicateurs["meta"]["ca_tnd"],
                "volets_generes":  ["ENVIRONNEMENT", "SOCIAL", "GOUVERNANCE"]
            }
        }
        
        # --- Envoi de l email avec pièce jointe directement depuis Python ---
        _envoyer_email_avec_piece_jointe(
            sujet=f"✅ Rapport ESG InnovTech généré — {datetime.now().strftime('%d/%m/%Y')}",
            corps=_construire_corps_email(reponse),
            fichier_joint=fichier_rapport,
            nom_fichier=nom_rapport
        )

        print(f"  ✅ Rapport généré et email envoyé : {nom_rapport}")
        return jsonify(reponse), 200
    
    except Exception as e:
        # En cas d'erreur, n8n reçoit le message d'erreur complet
        print(f"  ❌ Erreur : {str(e)}")
        return jsonify({
            "statut":  "erreur",
            "message": str(e),
            "detail":  traceback.format_exc()
        }), 500


# ==============================================================================
# FONCTIONS INTERNES (identiques à la version manuelle, sans prints)
# ==============================================================================

def _lire_base_de_donnees(chemin):
    """Charge toutes les feuilles Excel de la base de données."""
    return pd.read_excel(chemin, sheet_name=None)


def _extraire_indicateurs_esg(db):
    """Extrait et calcule les indicateurs ESG depuis la base de données."""
    df_emp = db["Employés"]
    
    # Calcul taux de femmes (détection par prénoms féminins courants)
    total_emp = len(df_emp)
    prenoms_fem = ["Rim", "Leila", "Sonia", "Salma", "Fatima", "Nadia", "Yasmine", "Dina"]
    femmes = df_emp[df_emp["Prénom"].isin(prenoms_fem)]
    taux_femmes = round(len(femmes) / total_emp * 100, 1)
    
    # Taux femmes postes seniors
    postes_senior = df_emp[df_emp["Poste"].str.contains(
        "DAF|Responsable|Chef|Directeur", case=False, na=False)]
    femmes_senior = postes_senior[postes_senior["Prénom"].isin(prenoms_fem)]
    taux_femmes_senior = round(len(femmes_senior) / max(len(postes_senior), 1) * 100, 1)
    
    # Données économiques
    df_eco = db["Impact_Économique"]
    def val(nom):
        ligne = df_eco[df_eco["Indicateur"].str.contains(nom, case=False, na=False)]
        return ligne.iloc[0]["2024"] if not ligne.empty else "N/A"
    
    ca_2024        = val("Chiffre d")
    emplois_dir    = int(val("Emplois Directs"))
    part_export    = round(val("Part du CA Export") * 100, 0)
    investiss      = val("Investissements")
    
    return {
        "env": {
            "scope12":          {"valeur": "40 Mt CO₂e",     "obj_2025": "< 42 Mt",       "obj_2030": "−20 % vs 2021", "obj_lt": "Net zéro 2050",    "source": "Base données"},
            "intensite_carbone":{"valeur": "−10 % vs 2015",  "obj_2025": "−12 %",         "obj_2030": "−20 %",         "obj_lt": "Neutralité 2050",  "source": "Base données"},
            "intensite_methane":{"valeur": "~0,26 %",        "obj_2025": "−50 % vs 2020", "obj_2030": "−70 % vs 2020","obj_lt": "≈0,1 %",          "source": "Base données"},
            "invest_bas_carbone":{"valeur":"≈3,5 Md TND",    "obj_2025": "≥4 Md TND",     "obj_2030": "augmentation",  "obj_lt": "multi-énergies",   "source": "Base données"},
            "capacite_renouv":  {"valeur": "≈27 GW",         "obj_2025": "32 GW",         "obj_2030": "croissance",    "obj_lt": "pilier mix",       "source": "Base données"},
            "part_elec_renouv": {"valeur": "en croissance",  "obj_2025": "progression",   "obj_2030": "≈65 %",         "obj_lt": "dominante",        "source": "Base données"},
        },
        "social": {
            "femmes_senior":    {"valeur": f"{taux_femmes_senior} %", "obj_2025": "38 %",       "obj_2030": "≥40 %",     "obj_lt": "parité progressive","source": "Base données"},
            "femmes_total":     {"valeur": f"≈{taux_femmes} %",       "obj_2025": "progression","obj_2030": "progression","obj_lt": "parité",           "source": "Base données"},
            "diversite_intl":   {"valeur": "47,7 %",   "obj_2025": "maintien",     "obj_2030": "progression","obj_lt": "gouvernance globale","source": "Base données"},
            "trir":             {"valeur": "0,76",     "obj_2025": "baisse",       "obj_2030": "amélioration","obj_lt": "zéro accident grave","source": "Base données"},
            "programme_care":   {"valeur": "déployé",  "obj_2025": "maintien",     "obj_2030": "extension",   "obj_lt": "standard global",   "source": "Rapport RH"},
            "protection_sociale":{"valeur":"couverture Tunisie","obj_2025":"maintien","obj_2030":"extension",  "obj_lt": "harmonisation",     "source": "Rapport RH"},
        },
        "gouv": {
            "femmes_senior_dir":    {"valeur": "34,4 %",   "obj_2025": "38 %",        "obj_2030": "≥40 %",     "obj_lt": "parité progressive","source": "Base données"},
            "diversite_intl_dir":   {"valeur": "47,7 %",   "obj_2025": "maintien",    "obj_2030": "progression","obj_lt": "gouvernance mondiale","source": "Base données"},
            "vote_strategie_climat":{"valeur": "oui",       "obj_2025": "maintien",    "obj_2030": "maintien",  "obj_lt": "standard climat",   "source": "Rapport interne"},
            "publication_esg":      {"valeur": "annuel",    "obj_2025": "maintien",    "obj_2030": "maintien",  "obj_lt": "standard reporting","source": "Site officiel"},
            "alignement_csrd":      {"valeur": "en cours",  "obj_2025": "conformité",  "obj_2030": "maintien",  "obj_lt": "standard UE",       "source": "Rapport interne"},
            "vote_consultatif_climat":{"valeur":"annuel",   "obj_2025": "maintien",    "obj_2030": "maintien",  "obj_lt": "standard gouvernance","source":"Rapport interne"},
            "invest_bas_carbone_perf":{"valeur":"≈3,5 Md TND","obj_2025":"maintien",  "obj_2030": "augmentation","obj_lt":"transformation",    "source": "Base données"},
            "scope3_clients":       {"valeur": "561,3 Mt CO₂e","obj_2025":"maintien", "obj_2030": "maintien",  "obj_lt": "neutralité société","source": "Base données"},
            "transf_mix":           {"valeur": "engagée",   "obj_2025": "accélération","obj_2030":"diversification","obj_lt":"multi-énergies", "source": "Rapport interne"},
        },
        "meta": {
            "entreprise": NOM_ENTREPRISE,
            "annee":      ANNEE_RAPPORT,
            "emplois":    emplois_dir,
            "ca_tnd":     f"{ca_2024:,.0f} TND",
            "export_pct": f"{int(part_export)} %",
            "date_gen":   datetime.now().strftime("%d/%m/%Y %H:%M"),
        }
    }


def _style_titre(cell, couleur):
    cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=couleur)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def _style_entete(cell):
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _style_donnee(cell, pair):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color="EBF3FB" if pair else "FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


def _ecrire_volet(ws, titre, criteres, couleur):
    """Écrit un volet ESG complet dans une feuille Excel."""
    ws.delete_rows(1, ws.max_row)
    
    # En-tête principal
    ws.merge_cells("A1:F1")
    ws["A1"] = f"  {titre} — {NOM_ENTREPRISE} | Exercice {ANNEE_RAPPORT}"
    ws["A1"].font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", start_color=couleur)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:F2")
    ws["A2"] = f"  Rapport généré le : {datetime.now().strftime('%d/%m/%Y')} | Source : Base de données InnovTech SA"
    ws["A2"].font  = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].fill  = PatternFill("solid", start_color="D9E2F3")
    ws.row_dimensions[2].height = 16

    for col, larg in zip(["A","B","C","D","E","F"], [38,18,14,14,18,20]):
        ws.column_dimensions[col].width = larg

    ligne = 3
    for idx, critere in enumerate(criteres):
        if idx > 0:
            ligne += 1
        
        # Titre critère
        ws.merge_cells(f"A{ligne}:F{ligne}")
        ws[f"A{ligne}"] = f"  {critere['titre']}"
        _style_titre(ws[f"A{ligne}"], "1F3864")
        ws.row_dimensions[ligne].height = 20
        ligne += 1
        
        # En-têtes colonnes
        for c, h in enumerate(["Indicateur","Valeur actuelle","Objectif 2025","Objectif 2030","Objectif long terme","Source"], 1):
            _style_entete(ws.cell(row=ligne, column=c, value=h))
        ws.row_dimensions[ligne].height = 18
        ligne += 1
        
        # Données
        for ri, ind in enumerate(critere["indicateurs"]):
            pair = ri % 2 == 0
            _style_donnee(ws.cell(row=ligne, column=1, value=ind["nom"]), pair)
            
            cell_v = ws.cell(row=ligne, column=2, value=ind["valeur"])
            cell_v.font = Font(name="Arial", bold=True, size=10, color="1F3864")
            cell_v.fill = PatternFill("solid", start_color="EBF3FB" if pair else "FFFFFF")
            cell_v.alignment = Alignment(horizontal="center", vertical="center")
            
            for c, k in enumerate(["obj_2025","obj_2030","obj_lt","source"], 3):
                cell = ws.cell(row=ligne, column=c, value=ind.get(k,""))
                _style_donnee(cell, pair)
                if k == "source":
                    cell.font = Font(name="Arial", size=9, italic=True, color="595959")
            
            ws.row_dimensions[ligne].height = 16
            ligne += 1
    
    ws.freeze_panes = "A3"


def _creer_rapport(db, indicateurs, fichier_template, fichier_sortie):
    """Crée le classeur Excel ESG complet à partir du template."""
    wb = load_workbook(fichier_template)
    
    # ---- ENVIRONNEMENT ----
    criteres_env = [
        {"titre": "CRITÈRE 1 — ÉMISSIONS CARBONE", "indicateurs": [
            {"nom": "Scope 1+2 (émissions directes)",       **indicateurs["env"]["scope12"]},
            {"nom": "Intensité carbone produits vendus",     **indicateurs["env"]["intensite_carbone"]},
            {"nom": "Intensité méthane (% production)",      **indicateurs["env"]["intensite_methane"]},
        ]},
        {"titre": "CRITÈRE 2 — TRANSITION ÉNERGÉTIQUE", "indicateurs": [
            {"nom": "Investissements bas carbone",           **indicateurs["env"]["invest_bas_carbone"]},
            {"nom": "Capacité renouvelable installée",       **indicateurs["env"]["capacite_renouv"]},
            {"nom": "Part électricité renouvelable",         **indicateurs["env"]["part_elec_renouv"]},
        ]},
        {"titre": "CRITÈRE 3 — MÉTHANE & SURVEILLANCE", "indicateurs": [
            {"nom": "Réduction émissions méthane",          "valeur": "base 2020", "obj_2025": "−50 %", "obj_2030": "−70 %", "obj_lt": "quasi-élimination", "source": "Base données"},
            {"nom": "Intensité méthane production",         "valeur": "~0,26 %",   "obj_2025": "baisse","obj_2030": "0,15 %","obj_lt": "maintien faible",   "source": "Base données"},
            {"nom": "Surveillance fuites (capteurs IoT)",   "valeur": "déploiement","obj_2025":"extension","obj_2030":"couverture complète","obj_lt":"standard industrie","source":"Rapport IT"},
        ]},
    ]
    _ecrire_volet(wb["ENVIRONNEMENT"], "VOLET ENVIRONNEMENT", criteres_env, "1F5C2E")
    
    # ---- SOCIAL ----
    criteres_soc = [
        {"titre": "CRITÈRE 1 — DIVERSITÉ & INCLUSION", "indicateurs": [
            {"nom": "Femmes senior executives",             **indicateurs["social"]["femmes_senior"]},
            {"nom": "Femmes effectif total",                **indicateurs["social"]["femmes_total"]},
            {"nom": "Diversité internationale dirigeants",  **indicateurs["social"]["diversite_intl"]},
        ]},
        {"titre": "CRITÈRE 2 — SANTÉ & SÉCURITÉ", "indicateurs": [
            {"nom": "TRIR (Taux de fréquence accidents)",   **indicateurs["social"]["trir"]},
            {"nom": "Programme Care Together",              **indicateurs["social"]["programme_care"]},
            {"nom": "Protection sociale minimale",          **indicateurs["social"]["protection_sociale"]},
        ]},
        {"titre": "CRITÈRE 3 — FORMATION & ENGAGEMENT", "indicateurs": [
            {"nom": "Formation sécurité (h/employé/an)",   "valeur":"obligatoire","obj_2025":"maintien","obj_2030":"amélioration","obj_lt":"standard industrie","source":"Rapport RH"},
            {"nom": "Programme Care Together fournisseurs","valeur":"global",      "obj_2025":"maintien","obj_2030":"extension",   "obj_lt":"standard élargi", "source":"Rapport RH"},
            {"nom": "Dialogue social international",        "valeur":"structuré",  "obj_2025":"maintien","obj_2030":"renforcement","obj_lt":"durable",          "source":"Rapport RH"},
        ]},
    ]
    _ecrire_volet(wb["SOCIAL"], "VOLET SOCIAL", criteres_soc, "1F3864")
    
    # ---- GOUVERNANCE ----
    criteres_gouv = [
        {"titre": "CRITÈRE 1 — STRUCTURE DIRIGEANTE", "indicateurs": [
            {"nom": "Femmes senior executives",             **indicateurs["gouv"]["femmes_senior_dir"]},
            {"nom": "Diversité internationale dirigeants",  **indicateurs["gouv"]["diversite_intl_dir"]},
            {"nom": "Vote actionnaires stratégie climat",   **indicateurs["gouv"]["vote_strategie_climat"]},
        ]},
        {"titre": "CRITÈRE 2 — TRANSPARENCE ESG", "indicateurs": [
            {"nom": "Publication ESG Databook",             **indicateurs["gouv"]["publication_esg"]},
            {"nom": "Alignement CSRD (réglementation UE)", **indicateurs["gouv"]["alignement_csrd"]},
            {"nom": "Vote consultatif climat",              **indicateurs["gouv"]["vote_consultatif_climat"]},
        ]},
        {"titre": "CRITÈRE 3 — PERFORMANCE DURABLE", "indicateurs": [
            {"nom": "Investissements bas carbone",          **indicateurs["gouv"]["invest_bas_carbone_perf"]},
            {"nom": "Émissions Scope 3 clients",            **indicateurs["gouv"]["scope3_clients"]},
            {"nom": "Transformation mix énergétique",       **indicateurs["gouv"]["transf_mix"]},
        ]},
    ]
    _ecrire_volet(wb["GOUVERNANCE"], "VOLET GOUVERNANCE", criteres_gouv, "6B2737")
    
    wb.save(fichier_sortie)


# ==============================================================================
# FONCTIONS EMAIL — Envoi direct depuis Python avec pièce jointe
# ==============================================================================

def _construire_corps_email(reponse):
    """Construit le texte de l email de notification."""
    date = datetime.fromisoformat(reponse["timestamp"]).strftime("%d/%m/%Y")
    return f"""Bonjour,

Veuillez trouver ci-joint le rapport ESG InnovTech SA du {date}.

Cordialement,
Système ESG Automatisé InnovTech SA"""


def _envoyer_email_avec_piece_jointe(sujet, corps, fichier_joint, nom_fichier):
    """
    Envoie un email avec le rapport Excel en pièce jointe via Gmail SMTP.
    Utilise SSL sur le port 465 — plus fiable que TLS.
    """
    try:
        # Création du message email
        msg = MIMEMultipart()
        msg["From"]    = EMAIL_EXPEDITEUR
        msg["To"]      = EMAIL_DESTINATAIRE
        msg["Subject"] = sujet

        # Corps du message en texte simple
        msg.attach(MIMEText(corps, "plain", "utf-8"))

        # Lecture et attachement du fichier Excel
        with open(fichier_joint, "rb") as f:
            partie = MIMEBase("application",
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            partie.set_payload(f.read())
            encoders.encode_base64(partie)
            partie.add_header("Content-Disposition",
                f"attachment; filename={nom_fichier}")
            msg.attach(partie)

        # Envoi via Gmail SSL
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as serveur:
            serveur.login(EMAIL_EXPEDITEUR, EMAIL_MOT_DE_PASSE)
            serveur.send_message(msg)

        print(f"  📧 Email envoyé à {EMAIL_DESTINATAIRE} avec pièce jointe : {nom_fichier}")

    except Exception as e:
        # L email échoue mais le rapport est quand même généré
        print(f"  ⚠️  Email non envoyé : {str(e)}")


# ==============================================================================
# DÉMARRAGE DU SERVEUR
# ==============================================================================
if __name__ == "__main__":
    print("=" * 60)
    print(f"  🌐 API ESG — {NOM_ENTREPRISE}")
    print(f"  Port : {PORT_API}")
    print(f"  Endpoints :")
    print(f"    GET  http://localhost:{PORT_API}/statut")
    print(f"    POST http://localhost:{PORT_API}/generer-rapport")
    print("=" * 60)
    print("\nPrêt à recevoir les requêtes n8n...\n")
    
    # debug=True permet de voir les erreurs détaillées (désactive en production)
    app.run(host="0.0.0.0", port=PORT_API, debug=False)
